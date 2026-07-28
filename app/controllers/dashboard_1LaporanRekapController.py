# controllers/dashboard_1LaporanRekapController.py
from flask import render_template, request, send_file
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image as XLImage
from app import db
from app.models.absensiModel import Absensi
from app.models.pegawaiModel import Pegawai
from app.models.kalenderModel import MfKalender
from app.models.dinasLuarModel import DinasLuar
from app.models.unitKerjaModel import MfUnitKerja


def laporan_cetak_daftar_lembur_umum():
    """
    Render halaman Laporan Cetak Daftar Lembur Umum.
    """
    return render_template('pages/dashboard_1/Laporan Cetak Daftar Lembur Umum.html')


def laporan_rekap_absensi_all():
    """
    Render halaman Laporan Rekap Absensi All.
    Unit Kerja dropdown diisi dari tabel MF_UNIT_KERJA (server-side render),
    bukan hardcode di HTML.
    """
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()

    return render_template(
        'pages/dashboard_1/Laporan Rekap Absensi All.html',
        unit_kerja_list=unit_kerja_list
    )

def export_rekap_absensi_all():
    unit_list = request.form.getlist('unit_kerja[]')
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'Unit kosong atau format tanggal salah'}, 400
    
    # Konversi unit_list ke integer
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    tampilkan_ket = request.form.get('kolom_keterangan') == 'tampilkan'
    nama_eselon3 = request.form.get('nama_eselon3', '')
    pangkat_eselon3 = request.form.get('pangkat_eselon3', '')
    petugas1 = request.form.get('petugas1', '')
    petugas2 = request.form.get('petugas2', '')

    # Cek tanggal server
    tgl_server = datetime.now()
    if tgl_server.date() < tgl_awal.date():
        return {'error': 'Tgl server lebih kecil dari tanggal awal periode'}, 400
    if tgl_server.date() < tgl_akhir.date():
        tgl_akhir = tgl_server

    # Query — JOIN VIA NIP, BUKAN FINGER_ID
    q = (
        db.session.query(Absensi, Pegawai)
        .join(Pegawai, Absensi.NIP == Pegawai.NIP)  # ✅ JOIN VIA NIP
        .join(MfKalender, Absensi.TGL_KERJA == MfKalender.TGL_KERJA)
        .filter(Absensi.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(MfKalender.IS_LIBUR == 'N')
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))  # ✅ pakai integer
    )
    rows = q.all()
    df_absensi = pd.DataFrame([{**a.__dict__, **p.__dict__} for a, p in rows]) if rows else pd.DataFrame()

    if df_absensi.empty:
        return {'error': 'Record tidak ada atau kalender belum dibuat'}, 400

    # Agregasi — ganti 'tingkat_tlm' jadi 'TINGKAT_TLM' (sesuai kolom model)
    hasil = []
    for nip, grp in df_absensi.groupby('NIP'):
        hasil.append({
            'nip': nip,
            'nama': grp['NAMA'].iloc[0],
            'tlm1': (grp['TINGKAT_TLM'] == 'TLM-1').sum(),
            'tlm2': (grp['TINGKAT_TLM'] == 'TLM-2').sum(),
            'tlm3': (grp['TINGKAT_TLM'] == 'TLM-3').sum(),
            'tlm4': (grp['TINGKAT_TLM'] == 'TLM-4').sum(),
            'psw1': (grp['TINGKAT_PSW'] == 'PSW-1').sum(),
            'psw2': (grp['TINGKAT_PSW'] == 'PSW-2').sum(),
            'psw3': (grp['TINGKAT_PSW'] == 'PSW-3').sum(),
            'psw4': (grp['TINGKAT_PSW'] == 'PSW-4').sum(),
            'dl': (grp['TRANSAKSI_IN'] == 'DinasLuar').sum(),
            'cuti': ((grp['TRANSAKSI_IN'] == 'Cuti') & (grp['TINGKAT_TLM'] == 'CT')).sum(),
            'cb1': ((grp['TRANSAKSI_IN'] == 'Cuti') & (grp['TINGKAT_TLM'] == 'CB-1')).sum(),
            'cb2': ((grp['TRANSAKSI_IN'] == 'Cuti') & (grp['TINGKAT_TLM'] == 'CB-2')).sum(),
            'cb3': ((grp['TRANSAKSI_IN'] == 'Cuti') & (grp['TINGKAT_TLM'] == 'CB-3')).sum(),
            'capm2': ((grp['TRANSAKSI_IN'] == 'Cuti') & (grp['TINGKAT_TLM'] == 'CAP-M2')).sum(),
            'cap': ((grp['TRANSAKSI_IN'] == 'Cuti') & (grp['TINGKAT_TLM'] == 'CAP')).sum(),
            'sakit': ((grp['TRANSAKSI_IN'] == 'Sakit') & (grp['TINGKAT_TLM'] == 'S-1')).sum(),
            'sakit2': ((grp['TRANSAKSI_IN'] == 'Sakit') & (grp['TINGKAT_TLM'] == 'S-2')).sum(),
            'sakit3': ((grp['TRANSAKSI_IN'] == 'Sakit') & (grp['TINGKAT_TLM'] == 'S-3')).sum(),
            'sakit4': ((grp['TRANSAKSI_IN'] == 'Sakit') & (grp['TINGKAT_TLM'] == 'S-4')).sum(),
            'sakit5': ((grp['TRANSAKSI_IN'] == 'Sakit') & (grp['TINGKAT_TLM'] == 'S-5')).sum(),
            'alpa': ((grp['TRANSAKSI_IN'] == 'Alpa') & (grp['PENDUKUNG_IN'] == 'Y')).sum(),
            'alpa_tanpa_ket': ((grp['TRANSAKSI_IN'] == 'Alpa') & (grp['PENDUKUNG_IN'] == 'N')).sum(),
        })
    df_hasil = pd.DataFrame(hasil)

    # 4. (Opsional) keterangan dinas luar/cuti kalau tampilkan_ket True
    if tampilkan_ket:
        dl_rows = (
            DinasLuar.query
            .filter(DinasLuar.TGL_AWAL_DINAS_LUAR <= tgl_akhir)
            .filter(DinasLuar.TGL_AKHIR_DINAS_LUAR >= tgl_awal)
            .all()
        )
        # susun jadi dict {nip: "1. ... \n2. ..."} sesuai format lama

    # 5. Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar"
    ws.sheet_properties.tabColor = "FF7B00"

    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except FileNotFoundError:
        pass

    ws.merge_cells('D2:AL2')
    ws['D2'] = 'Laporan Rekap Daftar Hadir Pegawai'
    ws['D2'].font = Font(bold=True)
    ws['D2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('D3:AL3')
    ws['D3'] = f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}"
    ws['D3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('D4:AL4')
    ws['D4'] = f"Unit : {', '.join(unit_list)}"
    ws['D4'].alignment = Alignment(horizontal='center')

    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    header_row = 5
    headers = ['No', 'Nama', 'TLM1', 'TLM2', 'TLM3', 'TLM4', 'Cuti', 'Sakit', 'Alpa']  # lengkapi sesuai kebutuhan
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=col, value=h)
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

    row = header_row + 1
    for i, r in enumerate(df_hasil.to_dict('records'), start=1):
        ws.cell(row=row, column=2, value=i).border = border
        ws.cell(row=row, column=3, value=f"{r['nip']}\n{r['nama']}").border = border
        ws.cell(row=row, column=4, value=r['tlm1'] or None).border = border
        ws.cell(row=row, column=5, value=r['tlm2'] or None).border = border
        ws.cell(row=row, column=6, value=r['tlm3'] or None).border = border
        ws.cell(row=row, column=7, value=r['tlm4'] or None).border = border
        ws.cell(row=row, column=8, value=r['cuti'] or None).border = border
        ws.cell(row=row, column=9, value=r['sakit'] or None).border = border
        ws.cell(row=row, column=10, value=r['alpa'] or None).border = border
        row += 1

    row += 2
    ws.cell(row=row, column=4, value='Mengetahui,')
    ws.cell(row=row+1, column=4, value='Pejabat Eselon III')
    ws.cell(row=row+4, column=4, value=nama_eselon3).font = Font(underline='single')
    ws.cell(row=row+5, column=4, value=pangkat_eselon3)

    ws.cell(row=row, column=32, value=f"Surabaya, {tgl_akhir:%d %B %Y}")
    ws.cell(row=row+1, column=32, value='Petugas Pengelola Daftar Hadir')
    ws.cell(row=row+3, column=32, value=f"1. {petugas1}")
    ws.cell(row=row+5, column=32, value=f"2. {petugas2}")

    # 6. Stream sebagai download, bukan simpan ke disk
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Laporan_Rekap_Daftar_Hadir_Peg_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def laporan_rekap_absensi_individu():
    """
    Render halaman Laporan Rekap Absensi Individu.
    """
    return render_template('pages/dashboard_1/Laporan Rekap Absensi Individu.html')

def export_rekap_absensi_individu():
    """
    Export Laporan Rekap Absensi Individu (per pegawai, detail per hari).
    Mirip dengan FillRekapAbsensiPerson di VB.NET.
    """
    nip_list = request.form.getlist('nip[]')  # ['198501232009122002', ...]
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not nip_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'NIP atau tanggal kosong'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    
    # Cek tanggal server
    tgl_server = datetime.now()
    if tgl_server.date() < tgl_awal.date():
        return {'error': 'Tgl server lebih kecil dari tanggal awal periode'}, 400
    if tgl_server.date() < tgl_akhir.date():
        tgl_akhir = tgl_server
    
    # 1. Ambil data kalender (hari kerja saja)
    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(MfKalender.IS_LIBUR == 'N')
        .order_by(MfKalender.TGL_KERJA.asc())
        .all()
    )
    
    if not kalender_rows:
        return {'error': 'Tidak ada hari kerja dalam periode tersebut'}, 400
    
    # 2. Ambil data absensi untuk NIP yang dipilih
    absensi_rows = (
        db.session.query(Absensi, Pegawai)
        .join(Pegawai, Absensi.NIP == Pegawai.NIP)
        .filter(Absensi.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(Absensi.NIP.in_(nip_list))
        .order_by(Pegawai.NAMA, Absensi.TGL_KERJA)
        .all()
    )
    
    # 3. Ambil data pegawai
    pegawai_rows = (
        Pegawai.query
        .filter(Pegawai.NIP.in_(nip_list))
        .order_by(Pegawai.NAMA)
        .all()
    )
    
    if not pegawai_rows:
        return {'error': 'Pegawai tidak ditemukan'}, 400
    
    # 4. Build data per pegawai per tanggal
    # Struktur: {nip: {nama: ..., unit_kerja: ..., rows: [{tgl, tlm, kategori_tlm, psw, kategori_psw, cuti, dl, sakit, sakit_a, alpa, alpa_a, ket}]}}
    absensi_dict = {}
    for a, p in absensi_rows:
        tgl_key = a.TGL_KERJA.strftime('%Y-%m-%d') if a.TGL_KERJA else None
        if a.NIP not in absensi_dict:
            absensi_dict[a.NIP] = {}
        absensi_dict[a.NIP][tgl_key] = a
    
    # 5. Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Individu"
    ws.sheet_properties.tabColor = "FF7B00"
    
    # Setup printer
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Header
    ws.merge_cells('D2:N2')
    ws['D2'] = 'Laporan Rekap Daftar Hadir Pegawai'
    ws['D2'].font = Font(bold=True, size=12)
    ws['D2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D3:N3')
    ws['D3'] = f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}"
    ws['D3'].alignment = Alignment(horizontal='center')
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Kolom header
    headers = ['No.', 'Tanggal', 'TLM\n(menit)', 'Kategori\nTLM', 'PSW\n(menit)', 
               'Kategori\nPSW', 'Cuti\n(hari)', 'Dinas\nLuar', 'Sakit\nDokter', 
               'Sakit\ntnp dr', 'Tdk Hadir\ndgn Izin', 'Tdk Hadir\nTanpa Ket', 'Keterangan']
    
    header_row = 5
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=col, value=h)
        c.border = border
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set lebar kolom
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 9
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 9
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 25
    
    row = header_row + 1
    
    for pegawai in pegawai_rows:
        # Baris nama pegawai (merge)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        ws.cell(row=row, column=2, value=pegawai.NAMA).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        for col in range(2, 15):
            ws.cell(row=row, column=col).border = border
        row += 1
        
        # Detail per hari
        no = 0
        for kl in kalender_rows:
            no += 1
            tgl_str = kl.TGL_KERJA.strftime('%Y-%m-%d') if kl.TGL_KERJA else ''
            
            # Border
            for col in range(2, 15):
                ws.cell(row=row, column=col).border = border
            
            ws.cell(row=row, column=2, value=no).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row, column=3, value=kl.TGL_KERJA.strftime('%d-%m-%Y') if kl.TGL_KERJA else '').alignment = Alignment(horizontal='left', vertical='center')
            
            # Cek absensi untuk tanggal ini
            absensi = absensi_dict.get(pegawai.NIP, {}).get(tgl_str)
            
            if absensi:
                ws.cell(row=row, column=4, value=absensi.AWAL_TLM or 0).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row, column=5, value=absensi.TINGKAT_TLM or '').alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=6, value=absensi.TOTAL_PSW or 0).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row, column=7, value=absensi.TINGKAT_PSW or '').alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=14, value=absensi.KET_IN or '').alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Kategori transaksi
                transaksi = (absensi.TRANSAKSI_IN or '').upper()
                pendukung = (absensi.PENDUKUNG_IN or '').upper()
                
                if transaksi == 'CUTI':
                    ws.cell(row=row, column=8, value=1)
                elif transaksi == 'DINASLUAR':
                    ws.cell(row=row, column=9, value=1)
                elif transaksi == 'SAKIT':
                    if pendukung == 'Y':
                        ws.cell(row=row, column=10, value=1)
                    else:
                        ws.cell(row=row, column=11, value=1)
                elif transaksi == 'ALPA':
                    if pendukung == 'Y':
                        ws.cell(row=row, column=12, value=1)
                    else:
                        ws.cell(row=row, column=13, value=1)
                elif transaksi == 'IJIN':
                    if pendukung == 'Y':
                        ws.cell(row=row, column=12, value=1)
                    else:
                        ws.cell(row=row, column=13, value=1)
            else:
                # Tidak ada record = Alpa tanpa keterangan
                ws.cell(row=row, column=13, value=1)
            
            # Alignment untuk kolom angka
            for col in [8, 9, 10, 11, 12, 13]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        row += 1  # Spasi antar pegawai
    
    # Keterangan di bawah
    row += 1
    ws.cell(row=row, column=2, value='Keterangan:').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=2, value='TLM')
    ws.cell(row=row, column=4, value=': Terlambat Masuk')
    row += 1
    ws.cell(row=row, column=2, value='PSW')
    ws.cell(row=row, column=4, value=': Pulang Sebelum Waktu')
    row += 1
    ws.cell(row=row, column=2, value='TLM (-)')
    ws.cell(row=row, column=4, value=': Datang Lebih Awal')
    
    # Stream download
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Laporan_Rekap_Daftar_Hadir_Per_Pegawai_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def search_pegawai_by_name():
    """
    API untuk search pegawai berdasarkan nama (untuk dropdown autocomplete).
    """
    keyword = request.args.get('keyword', '').strip()
    if len(keyword) < 2:
        return {'data': []}
    
    pegawai_list = (
        Pegawai.query
        .filter(Pegawai.NAMA.ilike(f'%{keyword}%'))
        .order_by(Pegawai.NAMA.asc())
        .limit(15)
        .all()
    )
    
    return {
        'data': [
            {
                'nip': p.NIP,
                'nama': p.NAMA,
                'jabatan': p.JABATAN,
            }
            for p in pegawai_list
        ]
    }

def laporan_rekap_absensi_log_finger():
    """
    Render halaman Laporan Rekap Absensi Log Finger.
    """
    return render_template('pages/dashboard_1/Laporan Rekap Absensi Log Finger.html')


def laporan_rekap_clock_exception():
    """
    Render halaman Laporan Rekap Clock Exception.
    """
    return render_template('pages/dashboard_1/Laporan Rekap Clock Exception.html')


def laporan_rekap_ketidakhadiran_pegawai():
    """
    Render halaman Laporan Rekap Ketidakhadiran Pegawai.
    """
    return render_template('pages/dashboard_1/Laporan Rekap Ketidakhadiran Pegawai.html')


def laporan_rekap_pelanggaran_disiplin():
    """
    Render halaman Laporan Rekap Pelanggaran Disiplin.
    Catatan: logika dasarnya kemungkinan mirip _get_data_pelanggaran() di
    dashboard_1HomeController.py, tapi dalam bentuk laporan rekap (bisa filter periode).
    """
    return render_template('pages/dashboard_1/Laporan Rekap Pelanggaran Disiplin.html')


def laporan_rekap_uang_makan():
    """
    Render halaman Laporan Rekap Uang Makan.
    """
    return render_template('pages/dashboard_1/Laporan Rekap Uang Makan.html')

def laporan_rekap_tunjangan_kinerja():
    """
    Render halaman Laporan Rekap Tunjangan Kinerja.
    """
    return render_template('pages/dashboard_1/Laporan Rincian Pembayaran Tunjangan.html')